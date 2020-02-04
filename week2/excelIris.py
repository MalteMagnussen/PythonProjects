import openpyxl

# created from: https://en.wikipedia.org/wiki/Iris_flower_data_set#Data_set
filename = './iris_data.xlsx'
wb = openpyxl.load_workbook(filename)
print(wb.sheetnames)

# $ python excelIris.py
# ["Fisher's Iris Data"]

# The openpyxl.load_workbook() function takes in the filename and
# returns a value of the workbook data type.
# This Workbook object represents the Excel file, a bit like how a File object represents an opened text file.
# Remember that iris_data.xlsx needs to be in the current working direc- tory in order for you to work with it.
# You can find out what the current working directory is by importing os and using os.getcwd(),
# and you can change the current working directory using os.chdir().

# Getting Sheets from the Workbook
# Each sheet is represented by a Worksheet object,
# which you can obtain by passing the sheet name string to the get_sheet_by_name() workbook method.
# Finally, you can call the get_active_sheet() method of a Workbook object to get the workbook’s active sheet.
# The active sheet is the sheet that’s on top when the workbook is opened in Excel.
# Once you have the Worksheet object, you can get its name from the title attribute.

sheet = wb["Fisher's Iris Data"]
print(sheet.title)
print(wb.active)


# ______________________________________________________
zelle = sheet['B1']
print(zelle)
print((zelle.column, zelle.row, zelle.value))

print(sheet.cell(row=1, column=2) == zelle)
print(sheet.cell(row=1, column=2).value)


for idx in range(1, 8, 2):
    print(idx, sheet.cell(row=idx, column=2).value)

# <Cell "Fisher's Iris Data".B1>
# (2, 1, 'Sepal width')
# True
# Sepal width
# 1 Sepal width
# 3 3
# 5 3.1
# 7 3.9
# _______________________________________________________

print(tuple(sheet['A1':'D4']))
for rowOfCellObjects in sheet['A1':'D4']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('---------')

# _______________________________________________________

for cellObj in list(sheet.columns)[1]:
    print(cellObj.value)
